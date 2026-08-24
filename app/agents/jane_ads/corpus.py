"""
Jane + Ads — corpus ingestion (ASC-SPEC-01 v2 §4).

Reads the Records sheet (header row 3, data from row 4) and maps each row to a
`Strategy`. The workbook uses human-readable labels; storage uses the spec's
canonical snake_case. This module owns that translation, so renaming a dropdown
label never silently rewrites stored data.

Two rules the spec states plainly and this module enforces:

  · EX-* rows are illustrations, not corpus records — skipped (§4.2). Note EX-04
    carries "Does not transfer"; it is an example of the verdict, not a record.
  · Rejected records DO import and are retained for anti-duplication (§17.3);
    they simply never retrieve.

Ingestion cannot approve. The store raises IngestionCannotApprove if a row
arrives as approved (§1.1, ENG fixture 12) — a sheet marked Approved by hand does
not get to bypass human review.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Iterable, Optional

from .backfill import derive_consumed_by
from .entities import (
    ConversionLocation,
    EvidenceGrade,
    MarketOrigin,
    PooledAccountSafety,
    SalesCycle,
    Strategy,
    StrategyCategory,
    StrategyPlatform,
    StrategyStatus,
    TransferVerdict,
)
from .store import IngestionCannotApprove, StrategyStore

HEADER_ROW = 3

# ── Workbook label -> canonical value ────────────────────────────────────────
CATEGORY_LABELS = {
    "Offer & Positioning": StrategyCategory.OFFER_POSITIONING,
    "Audience Construction": StrategyCategory.AUDIENCE_CONSTRUCTION,
    "Creative Formats & Hooks": StrategyCategory.CREATIVE_FORMATS,
    "Copy Angles": StrategyCategory.COPY_ANGLES,
    "Budget, Pacing & Timing": StrategyCategory.BUDGET_PACING,
    "Micro-Budget Testing": StrategyCategory.MICRO_BUDGET_TESTING,
    "Conversion Mechanics": StrategyCategory.CONVERSION_MECHANICS,
    "Retargeting & Sequencing": StrategyCategory.RETARGETING,
    "Platform Mechanics": StrategyCategory.PLATFORM_MECHANICS,
    "Diagnostics & Troubleshooting": StrategyCategory.DIAGNOSTICS,
}

PLATFORM_LABELS = {
    "Meta (FB/IG)": StrategyPlatform.META,
    "TikTok": StrategyPlatform.TIKTOK,
    "Google": StrategyPlatform.GOOGLE,
    "LinkedIn": StrategyPlatform.LINKEDIN,
    "Snapchat": StrategyPlatform.SNAPCHAT,
    "WhatsApp": StrategyPlatform.WHATSAPP,
    "Cross-platform": StrategyPlatform.CROSS_PLATFORM,
}

SALES_CYCLE_LABELS = {
    "Same day": SalesCycle.SAME_DAY,
    "1-7 days": SalesCycle.ONE_TO_SEVEN_DAYS,
    "1-4 weeks": SalesCycle.ONE_TO_FOUR_WEEKS,
    "Over a month": SalesCycle.OVER_A_MONTH,
    "Not applicable": SalesCycle.NOT_APPLICABLE,
}

MARKET_ORIGIN_LABELS = {
    "Nigeria": MarketOrigin.NIGERIA,
    "Nigeria (desk research)": MarketOrigin.NIGERIA_DESK_RESEARCH,
    "Africa (other)": MarketOrigin.AFRICA_OTHER,
    "US": MarketOrigin.US,
    "UK/EU": MarketOrigin.UK_EU,
    "Asia": MarketOrigin.ASIA,
    "Latin America": MarketOrigin.LATIN_AMERICA,
    "Global/Unspecified": MarketOrigin.GLOBAL_UNSPECIFIED,
}

VERDICT_LABELS = {
    "Applies as-is": TransferVerdict.APPLIES_AS_IS,
    "Applies with modification": TransferVerdict.APPLIES_WITH_MODIFICATION,
    "Does not transfer": TransferVerdict.DOES_NOT_TRANSFER,
}

STATUS_LABELS = {
    "Draft": StrategyStatus.DRAFT,
    "In review": StrategyStatus.IN_REVIEW,
    "Approved": StrategyStatus.APPROVED,
    "Rejected": StrategyStatus.REJECTED,
}

EXAMPLE_LABEL = "EXAMPLE"


@dataclass
class RowError:
    row: int
    strategy_id: str
    reason: str


@dataclass
class ImportReport:
    imported: int = 0
    skipped_examples: int = 0
    errors: list[RowError] = field(default_factory=list)

    @property
    def total_seen(self) -> int:
        return self.imported + self.skipped_examples + len(self.errors)

    def summary(self) -> str:
        return (
            f"{self.total_seen} rows read · {self.imported} imported · "
            f"{self.skipped_examples} EXAMPLE rows skipped · {len(self.errors)} rejected"
        )


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return None if v in ("", "None", "-", "N/A") else v
    return value


def _coerce_budget(value: Any) -> Optional[float]:
    v = _clean(value)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    digits = str(v).replace("₦", "").replace(",", "").strip()
    try:
        return float(digits)
    except ValueError:
        raise ValueError(f"budget floor {value!r} is not a ₦/day number")


def _lookup(label: Any, table: dict, field_name: str):
    v = _clean(label)
    if v is None:
        raise ValueError(f"{field_name} is mandatory")
    if v not in table:
        raise ValueError(f"{field_name}: {v!r} is not a value the Lists tab defines")
    return table[v]


def _reason(exc: Exception) -> str:
    lines = [ln.strip() for ln in str(exc).splitlines() if ln.strip()]
    for ln in lines:
        if ln.startswith("Value error,"):
            return ln.split("[type=")[0].strip()
    for ln in lines[1:]:
        if not ln.startswith("For further information"):
            return ln.split("[type=")[0].strip()
    return lines[0] if lines else str(exc)


def row_to_strategy(row: dict[str, Any]) -> Strategy:
    """One sheet row -> one Strategy at version 1.

    v2 fields other than consumed_by are NOT inferred here — see backfill.py for
    why guessing them is more dangerous than leaving them at their fail-closed
    defaults (pooled_account_safe=unknown blocks retrieval until a human reviews).
    """
    category = _lookup(row.get("Category"), CATEGORY_LABELS, "Category")
    grade_label = str(_clean(row.get("Evidence Grade")) or "")
    if not grade_label or grade_label[0] not in "ABCD":
        raise ValueError(f"Evidence Grade: {grade_label!r} is not A/B/C/D")

    return Strategy(
        strategy_id=str(_clean(row.get("ID")) or ""),
        version=1,
        status=STATUS_LABELS.get(str(_clean(row.get("Status")) or ""), StrategyStatus.DRAFT),
        category=category,
        claim=_clean(row.get("Claim")) or "",
        mechanism=_clean(row.get("Mechanism — why it works")) or "",
        evidence_grade=EvidenceGrade(grade_label[0]),
        market_origin=_lookup(row.get("Market Origin"), MARKET_ORIGIN_LABELS, "Market Origin"),
        transfer_verdict=_lookup(row.get("Transfer Verdict"), VERDICT_LABELS, "Transfer Verdict"),
        modification_required=_clean(row.get("Modification Required")),
        business_types=[b] if (b := _clean(row.get("Business Type"))) else [],
        budget_floor_ngn_daily=_coerce_budget(row.get("Budget Floor (₦/day)")),
        platforms=[_lookup(row.get("Platform"), PLATFORM_LABELS, "Platform")],
        funnel_stages=[f] if (f := _clean(row.get("Funnel Stage"))) else [],
        sales_cycle=_lookup(row.get("Sales Cycle"), SALES_CYCLE_LABELS, "Sales Cycle"),
        consumed_by=derive_consumed_by(category),
        # Phase 0 columns. Read from the sheet when present; absent stays
        # fail-closed (empty / unknown) rather than being inferred.
        conversion_location=(
            [ConversionLocation(_clean(row.get("Conversion Location")))]
            if _clean(row.get("Conversion Location")) else []
        ),
        pooled_account_safe=PooledAccountSafety(
            _clean(row.get("Pooled Account Safe")) or PooledAccountSafety.UNKNOWN.value
        ),
        source_type=_clean(row.get("Source Type")),
        source_reference=_clean(row.get("Source Link")),
        source_published_at=_clean(row.get("Source Date")),
        ingested_by=_clean(row.get("Seeded By")) or "import",
    )


async def import_rows(rows: Iterable[dict[str, Any]], store: StrategyStore) -> ImportReport:
    report = ImportReport()
    for offset, row in enumerate(rows):
        sheet_row = HEADER_ROW + 1 + offset
        raw_id = str(_clean(row.get("ID")) or "")
        if not raw_id:
            continue
        if str(_clean(row.get("Status")) or "") == EXAMPLE_LABEL or raw_id.startswith("EX-"):
            report.skipped_examples += 1
            continue
        try:
            strategy = row_to_strategy(row)
            await store.ingest(strategy)
        except IngestionCannotApprove as exc:
            report.errors.append(RowError(sheet_row, raw_id, str(exc)))
            continue
        except Exception as exc:                  # noqa: BLE001 — reported, not raised
            report.errors.append(RowError(sheet_row, raw_id, _reason(exc)))
            continue
        report.imported += 1
    return report


def read_records_sheet(path: str, sheet_name: str = "Records") -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = [ws.cell(HEADER_ROW, i).value for i in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        row = {h: ws.cell(r, i + 1).value for i, h in enumerate(headers) if h}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows
